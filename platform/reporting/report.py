"""
report.py — Generate PDF and Excel inspection reports.

PDF: Jinja2 HTML template → WeasyPrint → .pdf
Excel: openpyxl multi-sheet workbook → .xlsx
JSON: Compatible with existing output/inspection_report.json format
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from ml.src.utils import get_logger

logger = get_logger("axalon.reporting")

_TEMPLATES_DIR = Path(__file__).parent / "templates"

# ── Fault type display name mapping ───────────────────────────────────────────
# Maps YOLO canonical class names → IEC 62446-3 Annex C pattern names
_FAULT_TYPE_DISPLAY = {
    "cell":               "Single Cell Hot Spot",
    "cell-multi":         "Multi-Cell Hot Spot",
    "module":             "Module Thermal Anomaly",
    "string":             "String / Module Open Circuit",
    "bypass-diode":       "Bypass Diode / Junction Box Failure",
    "offline-module":     "Offline Module (Open Circuit)",
    "vegetation-shading": "Vegetation Shading",
    "soiling":            "Soiling / Partial Shading",
    "short-circuit":      "Module Short Circuit",
    "hot-spot-low":       "Hot Spot (10–40 K)",
    "hot-spot-high":      "Hot Spot (>40 K)",
}

# Maps our internal severity → IEC CoA label for display
_SEVERITY_DISPLAY = {
    "CRITICAL": "CoA 3 — Safety-Relevant",
    "HIGH":     "CoA 2-3 — Thermal Abnormality",
    "MEDIUM":   "CoA 2 — Thermal Abnormality",
    "LOW":      "CoA 2 — Monitor",
}

# IEC CoA integer → short display label
_COA_DISPLAY = {
    1: "CoA 1",
    2: "CoA 2",
    3: "CoA 3",
}


def _fault_display(class_name: str) -> str:
    return _FAULT_TYPE_DISPLAY.get(class_name, class_name.replace("-", " ").title())


def compute_revenue_loss(detections: list[dict], economics: dict) -> float:
    """Estimate daily revenue loss (USD) from CRITICAL and HIGH faults.

    Counts unique known panels (excludes R?-C?) with CRITICAL or HIGH severity.
    Formula: affected_panels × panel_capacity_kw × peak_sun_hours × cost_per_kwh
    """
    cost = float(economics.get("cost_per_kwh_usd", 0.08))
    capacity = float(economics.get("panel_capacity_kw", 0.4))
    hours = float(economics.get("peak_sun_hours_per_day", 5.0))
    affected = {
        d["panel_id"] for d in detections
        if d.get("severity") in ("CRITICAL", "HIGH")
        and d.get("panel_id") not in (None, "R?-C?")
    }
    return round(len(affected) * capacity * hours * cost, 2)


def _severity_display(severity: str) -> str:
    return _SEVERITY_DISPLAY.get(severity, severity.title())


def _coa_display(coa: int | None) -> str:
    return _COA_DISPLAY.get(coa, "CoA 2") if coa is not None else "CoA 2"


# ── Panel ID parser ───────────────────────────────────────────────────────────
_PANEL_RE = re.compile(r"^R(\d+)-C(\d+)$", re.IGNORECASE)


def _parse_panel_id(panel_id: str) -> dict:
    """Split 'R3-C7' into structured fields. Returns N/A for unknown formats."""
    m = _PANEL_RE.match(panel_id or "")
    if m:
        return {
            "Block_No": "N/A",
            "Row_No":   m.group(1),
            "Table_No": "N/A",
            "Panel_No": m.group(2),
        }
    return {
        "Block_No": "N/A",
        "Row_No":   "N/A",
        "Table_No": "N/A",
        "Panel_No": panel_id or "N/A",
    }


def generate_json_report(
    batch_result: dict,
    output_path: str | Path,
    corrections: list[dict] | None = None,
) -> Path:
    """Write inspection result as JSON (compatible with existing report format)."""
    if corrections is not None:
        batch_result = {**batch_result, "corrections": corrections}
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(batch_result, indent=2, default=str))
    logger.info("JSON report saved: %s", output_path)
    return output_path


def generate_excel_report(
    batch_result: dict,
    output_path: str | Path,
    site_meta: dict | None = None,
) -> Path:
    """Generate multi-sheet Excel report matching reference defect-list format.

    Sheets:
        1. Summary       — Fault Type → Count pivot (reference style)
        2. <park_id>     — Full defect list: id/Block_No/Row_No/Table_No/Panel_No/
                           Fault_Type/Critical/Min_Temp/Max_Temp/Delta_Temp
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    site_meta = site_meta or {}
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    park_id = batch_result.get("park_id", "inspection")

    # Collect all detections across results
    all_dets = []
    for result in batch_result.get("results", []):
        for det in result.get("detections", []):
            all_dets.append({**det, "image_id": result.get("image_id", "")})
    # For single-image results (inspect_pair)
    for det in batch_result.get("detections", []):
        all_dets.append({**det, "image_id": batch_result.get("image_id", "")})

    # Build fault-type count pivot
    fault_counts: dict[str, int] = {}
    for det in all_dets:
        ft = _fault_display(det.get("class", ""))
        fault_counts[ft] = fault_counts.get(ft, 0) + 1

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_font   = Font(bold=True)
    title_font = Font(bold=True, size=12)
    thin       = Side(style="thin")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_fill  = PatternFill("solid", fgColor="D9D9D9")

    def _style_header_row(ws, row: int, col_count: int) -> None:
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font
            cell.border = border

    def _auto_col_width(ws) -> None:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    # Title block (matches reference style)
    site_name = site_meta.get("site_name") or park_id
    title_cell = ws_sum.cell(row=1, column=1, value=f"{site_name}_Defect Summary")
    title_cell.font = title_font

    ws_sum.cell(row=2, column=1, value="Row Labels").font = hdr_font
    ws_sum.cell(row=2, column=2, value="Count of Fault_Type").font = hdr_font
    ws_sum.cell(row=2, column=1).border = border
    ws_sum.cell(row=2, column=2).border = border

    for i, (ft, cnt) in enumerate(sorted(fault_counts.items()), start=3):
        ws_sum.cell(row=i, column=1, value=ft).border = border
        ws_sum.cell(row=i, column=2, value=cnt).border = border

    total_row = 3 + len(fault_counts)
    ws_sum.cell(row=total_row, column=1, value="Grand Total").font = hdr_font
    ws_sum.cell(row=total_row, column=1).fill = grey_fill
    ws_sum.cell(row=total_row, column=1).border = border
    ws_sum.cell(row=total_row, column=2, value=len(all_dets)).font = hdr_font
    ws_sum.cell(row=total_row, column=2).fill = grey_fill
    ws_sum.cell(row=total_row, column=2).border = border

    economics = {
        "cost_per_kwh_usd": site_meta.get("cost_per_kwh_usd", 0.08),
        "panel_capacity_kw": site_meta.get("panel_capacity_kw", 0.4),
        "peak_sun_hours_per_day": site_meta.get("peak_sun_hours_per_day", 5.0),
    }
    rev_loss = compute_revenue_loss(all_dets, economics)
    rev_row = total_row + 1
    ws_sum.cell(row=rev_row, column=1, value="Est. Daily Revenue at Risk (USD)").font = hdr_font
    ws_sum.cell(row=rev_row, column=2, value=f"${rev_loss:.2f}").font = hdr_font

    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 22

    # ── Sheet 2: Full Defect List ─────────────────────────────────────────────
    sheet_name = re.sub(r"[\\/*?\[\]:]", "_", park_id)[:31]
    ws_det = wb.create_sheet(sheet_name)

    det_headers = [
        "id", "Block_No", "Row_No", "Table_No", "Panel_No",
        "Fault_Type", "IEC_CoA", "Severity", "Min_Temp", "Max_Temp",
        "Delta_T_Measured_K", "Delta_T_Norm_1000Wm2", "Irradiance_Wm2",
        "Abnormality_Type", "Recommended_Action",
    ]
    for c, h in enumerate(det_headers, start=1):
        cell = ws_det.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = grey_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    _SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
    all_dets.sort(key=lambda d: _SEV_ORDER.get(d.get("severity", "LOW"), 99))

    for idx, det in enumerate(all_dets, start=1):
        loc = _parse_panel_id(det.get("panel_id", ""))
        delta_t = det.get("delta_t_measured")
        delta_t_norm = det.get("delta_t_normalized")
        irradiance = det.get("irradiance_wm2")
        row_data = [
            str(idx),
            loc["Block_No"],
            loc["Row_No"],
            loc["Table_No"],
            loc["Panel_No"],
            _fault_display(det.get("class", "")),
            _coa_display(det.get("iec_coa")),
            _severity_display(det.get("severity", "")),
            "",   # Min_Temp — populated when absolute temp data available
            "",   # Max_Temp
            f"{delta_t:.1f}" if delta_t is not None else "",
            f"{delta_t_norm:.1f}" if delta_t_norm is not None else "",
            f"{irradiance:.0f}" if irradiance is not None else "",
            det.get("abnormality_type", "extended"),
            det.get("iec_action", ""),
        ]
        for c, val in enumerate(row_data, start=1):
            cell = ws_det.cell(row=idx + 1, column=c, value=val)
            cell.border = border

    _auto_col_width(ws_det)

    wb.save(str(output_path))
    logger.info("Excel report saved: %s", output_path)
    return output_path


def generate_pdf_report(
    batch_result: dict,
    output_path: str | Path,
    site_meta: dict | None = None,
) -> Path:
    """Generate PDF inspection report using Jinja2 + WeasyPrint."""
    try:
        from jinja2 import Environment, FileSystemLoader
        from weasyprint import HTML
    except ImportError:
        logger.error("PDF requires: pip install jinja2 weasyprint")
        raise

    site_meta = site_meta or {}
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Collect all detections
    all_detections = []
    for result in batch_result.get("results", []):
        for det in result.get("detections", []):
            all_detections.append({**det, "image_id": result.get("image_id", "")})
    for det in batch_result.get("detections", []):
        all_detections.append({**det, "image_id": batch_result.get("image_id", "")})

    _SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
    all_detections.sort(key=lambda d: _SEV_ORDER.get(d.get("severity", "LOW"), 99))

    # Build fault-type count table for defect summary section
    fault_counts: dict[str, int] = {}
    for det in all_detections:
        ft = _fault_display(det.get("class", ""))
        fault_counts[ft] = fault_counts.get(ft, 0) + 1

    # Enrich detections with display names and IEC fields
    for det in all_detections:
        det["fault_display"]    = _fault_display(det.get("class", ""))
        det["severity_display"] = _severity_display(det.get("severity", ""))
        det["coa_display"]      = _coa_display(det.get("iec_coa"))
        loc = _parse_panel_id(det.get("panel_id", ""))
        det["row_no"]   = loc["Row_No"]
        det["panel_no"] = loc["Panel_No"]

    # IEC CoA summary counts
    coa_counts = {1: 0, 2: 0, 3: 0}
    for det in all_detections:
        coa = det.get("iec_coa", 2)
        if coa in coa_counts:
            coa_counts[coa] += 1

    irradiance_wm2 = site_meta.get("irradiance_wm2", "N/A")

    context = {
        # Site metadata (IEC §8 inspection report fields)
        "client":          site_meta.get("client", "N/A"),
        "site_name":       site_meta.get("site_name") or batch_result.get("park_id", "N/A"),
        "location":        site_meta.get("location", "N/A"),
        "lat":             site_meta.get("lat", "N/A"),
        "lon":             site_meta.get("lon", "N/A"),
        "capacity_mw":     site_meta.get("capacity_mw", "N/A"),
        "drone_model":     site_meta.get("drone_model", "N/A"),
        # IEC §8f — environmental conditions
        "irradiance_wm2":  irradiance_wm2,
        "wind_speed_bft":  site_meta.get("wind_speed_bft", "N/A"),
        "cloud_coverage_okta": site_meta.get("cloud_coverage_okta", "N/A"),
        "air_temp_c":      site_meta.get("air_temp_c", "N/A"),
        "inspection_time": site_meta.get("inspection_time", "N/A"),
        # IEC §8h — inspection procedure
        "inspection_level": site_meta.get("inspection_level", "Simplified"),
        "emissivity":       site_meta.get("emissivity", "0.85"),
        "t_refl":           site_meta.get("t_refl_c", "N/A"),
        # Inspection data
        "park_id":            batch_result.get("park_id", "N/A"),
        "flight_date":        batch_result.get("flight_date", "N/A"),
        "generated_at":       datetime.now().strftime("%d/%m/%Y %H:%M"),
        "total_images":       batch_result.get("total_images", 0),
        "total_detections":   batch_result.get("total_detections", len(all_detections)),
        "fault_counts":       sorted(fault_counts.items()),
        "detections":         all_detections[:200],
        # IEC CoA summary
        "coa_counts":         coa_counts,
        "standard_ref":       "IEC TS 62446-3:2017",
        # Revenue loss estimation
        "economics": {
            "cost_per_kwh_usd": site_meta.get("cost_per_kwh_usd", 0.08),
            "panel_capacity_kw": site_meta.get("panel_capacity_kw", 0.4),
            "peak_sun_hours_per_day": site_meta.get("peak_sun_hours_per_day", 5.0),
        },
        "revenue_loss_usd": compute_revenue_loss(all_detections, {
            "cost_per_kwh_usd": site_meta.get("cost_per_kwh_usd", 0.08),
            "panel_capacity_kw": site_meta.get("panel_capacity_kw", 0.4),
            "peak_sun_hours_per_day": site_meta.get("peak_sun_hours_per_day", 5.0),
        }),
    }

    env = Environment(loader=FileSystemLoader(str(_TEMPLATES_DIR)))
    template = env.get_template("report.html")
    html_content = template.render(**context)

    HTML(string=html_content, base_url=str(_TEMPLATES_DIR)).write_pdf(str(output_path))
    logger.info("PDF report saved: %s", output_path)
    return output_path
